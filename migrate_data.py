"""
Script de migracion: Excel 2024 -> Supabase
Carga todos los datos maestros automaticamente.
"""
import openpyxl
from supabase import create_client

# --- CONFIGURACION ---
SUPABASE_URL = "https://oeqnuhxbgqvovstzqcga.supabase.co"
SUPABASE_KEY = "sb_publishable_rUeJGhnDTZPkTt-zD0dfew_ThX6ZssA"
EXCEL_PATH = r"C:\Users\cruiz\OneDrive\Pynthon\Scripts\Sistema Enologico\Movimiento OC Insumos Enologicos 2024.xlsx"

db = create_client(SUPABASE_URL, SUPABASE_KEY)


def migrate_processes():
    print("Migrando procesos enologicos...")
    processes = [
        "Vendimia", "Levadura", "Enzima", "Correccion", "Correcciones",
        "Est. Proteica", "Est. Tartarica", "Clarificante", "Bentonita",
        "Madera", "Maderas", "Filtracion Tangencial", "Filtracion Placas",
        "Filtracion Borras", "Filtracion Envasado", "Higiene",
        "Higiene y Sanitizacion", "Microbiologia", "Envasado", "Mosto",
        "Mezclas y Estabilizacion", "Servicios Externos",
    ]
    for p in processes:
        try:
            db.table("winemaking_processes").insert({"name": p}).execute()
        except Exception:
            pass
    print(f"  {len(processes)} procesos")


def migrate_grape_varieties():
    print("Migrando cepas...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Precios"]

    grapes_raw = set()
    for row in ws.iter_rows(min_row=2, max_col=20, values_only=True):
        val = row[9]  # col J = Cepa
        if val and str(val).strip():
            grapes_raw.add(str(val).strip())
    wb.close()

    grape_names = {
        "CS": ("Cabernet Sauvignon", "Tinto"), "MR": ("Merlot", "Tinto"),
        "SY": ("Syrah", "Tinto"), "PN": ("Pinot Noir", "Tinto"),
        "CR": ("Carmenere", "Tinto"), "MB": ("Malbec", "Tinto"),
        "CG": ("Cabernet Gris", "Tinto"), "CF": ("Cabernet Franc", "Tinto"),
        "TT": ("Tinto (Generico)", "Tinto"), "TTO": ("Tinto Total", "Tinto"),
        "TG": ("Tinto Generico", "Tinto"), "TP": ("Tempranillo", "Tinto"),
        "PT": ("Petit Verdot", "Tinto"), "MO": ("Mourvedre", "Tinto"),
        "GR": ("Grenache", "Tinto"), "CA": ("Carignan", "Tinto"),
        "PA": ("Pais", "Tinto"), "CI": ("Cinsault", "Tinto"),
        "CH": ("Chardonnay", "Blanco"), "SB": ("Sauvignon Blanc", "Blanco"),
        "RS": ("Riesling", "Blanco"), "VG": ("Viognier", "Blanco"),
        "GW": ("Gewurztraminer", "Blanco"), "PG": ("Pinot Gris", "Blanco"),
        "SE": ("Semillon", "Blanco"), "MU": ("Muscat", "Blanco"),
        "TO": ("Torrontes", "Blanco"), "BL": ("Blanco (Generico)", "Blanco"),
        "RO": ("Rosado", "Rosado"), "BO": ("Borras", "Borras"),
    }

    count = 0
    for code in sorted(grapes_raw):
        name, wtype = grape_names.get(code, (code, "Tinto"))
        try:
            db.table("grape_varieties").insert({
                "code": code, "name": name, "wine_type": wtype
            }).execute()
            count += 1
        except Exception:
            pass
    print(f"  {count} cepas")


def migrate_product_lines():
    print("Migrando lineas de producto...")
    lines = [
        ("Entry Level", 1), ("Varietal", 2), ("Varietal 2", 3),
        ("Reserva", 4), ("Reserva 2", 5), ("Gran Reserva", 6),
        ("Gran Reserva 2", 7), ("Premium", 8), ("Premium 2", 9),
        ("Premium 3", 10), ("Icono", 11), ("Heredium", 12), ("Generico", 13),
    ]
    for name, order in lines:
        try:
            db.table("product_lines").insert({"name": name, "sort_order": order}).execute()
        except Exception:
            pass
    print(f"  {len(lines)} lineas")


def migrate_workers():
    print("Migrando operarios...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Precios"]

    workers = set()
    for row in ws.iter_rows(min_row=2, max_col=20, values_only=True):
        val = row[17]  # col R = Operador
        if val and str(val).strip():
            workers.add(str(val).strip())
    wb.close()

    # Tambien extraer de Salida de Insumos
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Salida de Insumos"]
    for row in ws.iter_rows(min_row=2, max_col=21, values_only=True):
        val = row[19]  # col T = Nombre Operario
        if val and str(val).strip():
            workers.add(str(val).strip())
    wb.close()

    count = 0
    for w in sorted(workers):
        try:
            db.table("workers").insert({"full_name": w}).execute()
            count += 1
        except Exception:
            pass
    print(f"  {count} operarios")


def migrate_suppliers():
    print("Migrando proveedores...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Base de Dato"]

    suppliers = set()
    for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
        val = row[4]  # col E = Proveedor
        if val and str(val).strip():
            suppliers.add(str(val).strip())
    wb.close()

    count = 0
    for s in sorted(suppliers):
        try:
            db.table("suppliers").insert({"name": s}).execute()
            count += 1
        except Exception:
            pass
    print(f"  {count} proveedores")


def migrate_supplies():
    print("Migrando insumos enologicos...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Precios"]

    count = 0
    for row in ws.iter_rows(min_row=2, max_col=7, values_only=True):
        name = row[0]  # col A = Insumo
        if not name or not str(name).strip():
            continue

        cost = row[1]  # col B = Costo Unitario
        currency_raw = row[2]  # col C = Moneda
        item_id = row[4]  # col E = ID
        code = row[5]  # col F = Codigo

        # Normalizar moneda
        currency = "CLP"
        if currency_raw:
            curr_str = str(currency_raw).strip().lower()
            if "euro" in curr_str:
                currency = "EUR"
            elif "dolar" in curr_str or "usd" in curr_str:
                currency = "USD"

        # Normalizar unidad desde el nombre o datos
        unit = "Kg"
        name_lower = str(name).lower()
        if any(x in name_lower for x in ["liquida", "liquido", "lts", "litro", "perac"]):
            unit = "Lts"
        elif any(x in name_lower for x in ["placa", "cilindro", "oenosteryl", "antiflor", "sanitas", "deacid"]):
            unit = "Unidad"

        data = {
            "name": str(name).strip(),
            "code": str(code).strip() if code else f"InsEnol {item_id}",
            "unit": unit,
            "currency": currency,
        }
        if cost and isinstance(cost, (int, float)):
            data["unit_cost"] = float(cost)

        try:
            db.table("supplies").insert(data).execute()
            count += 1
        except Exception as e:
            print(f"  Error con {name}: {e}")
    wb.close()
    print(f"  {count} insumos")


def migrate_tanks():
    print("Migrando cubas/tanques...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Salida de Insumos"]

    tanks = set()
    for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
        source = row[9]   # col J = Cuba Inicial
        dest = row[10]     # col K = Cuba Destino
        if source and str(source).strip():
            tanks.add(str(source).strip())
        if dest and str(dest).strip():
            tanks.add(str(dest).strip())
    wb.close()

    # Leer capacidades desde Precios
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Precios"]
    tank_caps = {}
    for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
        tk = row[14]   # col O = TK
        cap = row[15]  # col P = Capacidad
        if tk and cap:
            tank_caps[str(tk).strip()] = cap
    wb.close()

    count = 0
    for t in sorted(tanks, key=lambda x: (not x.isdigit(), x)):
        capacity = tank_caps.get(t)
        data = {"code": t}
        if capacity and isinstance(capacity, (int, float)):
            data["capacity_liters"] = int(capacity)
        try:
            db.table("tanks").insert(data).execute()
            count += 1
        except Exception:
            pass
    print(f"  {count} cubas/tanques")


def migrate_parametros():
    print("Migrando mapeo insumo-proceso...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["parametros"]

    # Obtener IDs de supplies y processes
    supplies = {s["name"]: s["id"] for s in db.table("supplies").select("id, name").execute().data}
    processes = {p["name"]: p["id"] for p in db.table("winemaking_processes").select("id, name").execute().data}

    count = 0
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        insumo = row[1] if len(row) > 1 else None  # col B
        proceso = row[2] if len(row) > 2 else None  # col C
        if not insumo or not proceso:
            continue

        supply_id = supplies.get(str(insumo).strip())
        process_id = processes.get(str(proceso).strip())

        if supply_id and process_id:
            try:
                db.table("supply_process_map").insert({
                    "supply_id": supply_id, "process_id": process_id
                }).execute()
                count += 1
            except Exception:
                pass
    wb.close()
    print(f"  {count} mapeos")


if __name__ == "__main__":
    print("=" * 50)
    print("MIGRACION DE DATOS: Excel 2024 -> Supabase")
    print("=" * 50)
    print()

    migrate_processes()
    migrate_grape_varieties()
    migrate_product_lines()
    migrate_workers()
    migrate_suppliers()
    migrate_supplies()
    migrate_tanks()
    migrate_parametros()

    print()
    print("=" * 50)
    print("MIGRACION COMPLETADA")
    print("=" * 50)
    print("Recarga la app de Streamlit para ver los datos.")
